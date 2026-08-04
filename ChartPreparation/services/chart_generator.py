from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import csv
from io import StringIO
import os
import sys
from pathlib import Path


def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    if getattr(sys, 'frozen', False):
        # Running as compiled exe
        base_path = sys._MEIPASS
    else:
        # Running as script
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def load_stock_symbols():
    """
    Load stock symbols from StockSymbols.csv
    Returns (success, symbols_set or error_message)
    """
    try:
        symbols_file = get_resource_path("StockSymbols.csv")
        
        if not os.path.exists(symbols_file):
            return False, "StockSymbols.csv file not found."
        
        with open(symbols_file, 'r', encoding='utf-8') as f:
            symbols = {line.strip() for line in f if line.strip()}
        
        if not symbols:
            return False, "StockSymbols.csv is empty."
        
        return True, symbols
        
    except Exception as e:
        return False, f"Error loading StockSymbols.csv: {str(e)}"


def parse_csv_with_flexible_columns(csv_content):
    """
    Parse CSV and return rows with normalized column names (stripped of spaces)
    Returns (success, data_dict or error_message)
    """
    try:
        csv_file = StringIO(csv_content)
        reader = csv.DictReader(csv_file)
        
        # Normalize column names by stripping spaces
        if reader.fieldnames:
            normalized_fieldnames = [col.strip() for col in reader.fieldnames]
            reader.fieldnames = normalized_fieldnames
        
        rows = list(reader)
        
        if len(rows) == 0:
            return False, "CSV file contains no data rows."
        
        return True, rows
        
    except Exception as e:
        return False, f"Error parsing CSV: {str(e)}"


def validate_csv_columns(rows):
    """
    Validate that required columns exist in the CSV
    """
    required_columns = ['SYMBOL', 'SERIES', 'PREV_CLOSE', 'HIGH_PRICE', 
                       'LOW_PRICE', 'CLOSE_PRICE', 'TTL_TRD_QNTY']
    
    if not rows:
        return False, "No data rows found."
    
    first_row = rows[0]
    missing_columns = [col for col in required_columns if col not in first_row]
    
    if missing_columns:
        return False, f"Missing required columns: {', '.join(missing_columns)}"
    
    return True, "All required columns present."


def filter_and_process_data(rows, stock_symbols):
    """
    Filter rows by stock symbols and process the data
    Returns list of processed row dictionaries
    """
    processed_rows = []
    
    for row in rows:
        symbol = row.get('SYMBOL', '').strip()
        
        # Filter by stock symbols
        if symbol not in stock_symbols:
            continue
        
        try:
            # Extract and convert numeric values
            close = float(row.get('CLOSE_PRICE', 0))
            high = float(row.get('HIGH_PRICE', 0))
            low = float(row.get('LOW_PRICE', 0))
            prv = float(row.get('PREV_CLOSE', 0))
            vol = float(row.get('TTL_TRD_QNTY', 0))
            series = row.get('SERIES', '').strip()
            
            # Calculate derived values
            # % = (CLOSE-PRV)*100/PRV
            percentage = ((close - prv) * 100 / prv) if prv != 0 else 0
            
            # PP = ABS((CLOSE+HIGH+LOW)/3)
            pp = abs((close + high + low) / 3)
            
            # R1 = ABS(PP*2-LOW)
            r1 = abs(pp * 2 - low)
            
            # S1 = ABS(PP*2-HIGH)
            s1 = abs(pp * 2 - high)
            
            # R2 = ABS(R1-S1+PP)
            r2 = abs(r1 - s1 + pp)
            
            # S2 = ABS(PP-(R1-S1))
            s2 = abs(pp - (r1 - s1))
            
            # R3 = ABS(HIGH+2*(PP-LOW))
            r3 = abs(high + 2 * (pp - low))
            
            # S3 = ABS(LOW-2*(HIGH-PP))
            s3 = abs(low - 2 * (high - pp))
            
            processed_rows.append({
                'SYMBOL': symbol,
                'CLOSE': close,
                'S1': s1,
                'S2': s2,
                'S3': s3,
                'PP': pp,
                'R1': r1,
                'R2': r2,
                'R3': r3,
                '%': percentage,
                'HIGH': high,
                'LOW': low,
                'PRV': prv,
                'VOL': vol,
                'SERIES': series
            })
            
        except (ValueError, TypeError) as e:
            # Skip rows with invalid numeric data
            continue
    
    return processed_rows


def create_excel_output(processed_rows, filename):
    """
    Create Excel file with processed data
    Returns (success, filepath or error_message)
    """
    try:
        # Sort processed rows by % column in descending order
        processed_rows_sorted = sorted(processed_rows, key=lambda x: x.get('%', 0), reverse=True)
        
        # Create workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Daily Chart"
        
        # Define column headers in the required order
        headers = ['SYMBOL', 'CLOSE', 'S1', 'S2', 'S3', 'PP', 
                  'R1', 'R2', 'R3', '%', 'HIGH', 'LOW', 'PRV', 'VOL', 'SERIES']
        
        # Write headers with formatting
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows (sorted by % descending)
        for row_idx, data_row in enumerate(processed_rows_sorted, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = data_row.get(header, '')
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Format numeric cells to 2 decimal places (except SYMBOL, SERIES, VOL)
                if header not in ['SYMBOL', 'SERIES', 'VOL'] and isinstance(value, (int, float)):
                    cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            sheet.column_dimensions[chr(64 + col_idx)].width = 12
        
        # Get Downloads folder path (cross-platform)
        downloads_folder = str(Path.home() / "Downloads")
        filepath = os.path.join(downloads_folder, filename)
        
        # Save workbook
        workbook.save(filepath)
        
        return True, filepath
        
    except Exception as e:
        return False, f"Error creating Excel file: {str(e)}"


# ==================== INDEX CHART FUNCTIONS ====================

def load_nifty_symbols():
    """
    Load index symbols from NiftySymbols.csv
    Returns (success, symbols_set or error_message)
    """
    try:
        symbols_file = get_resource_path("NiftySymbols.csv")
        
        if not os.path.exists(symbols_file):
            return False, "NiftySymbols.csv file not found."
        
        with open(symbols_file, 'r', encoding='utf-8') as f:
            symbols = {line.strip() for line in f if line.strip()}
        
        if not symbols:
            return False, "NiftySymbols.csv is empty."
        
        return True, symbols
        
    except Exception as e:
        return False, f"Error loading NiftySymbols.csv: {str(e)}"


def parse_index_csv(csv_content):
    """
    Parse Index CSV with flexible column handling
    Returns (success, data_dict or error_message)
    """
    try:
        # Remove any remaining BOM characters
        csv_content = csv_content.replace('\ufeff', '')
        
        csv_file = StringIO(csv_content)
        reader = csv.DictReader(csv_file)
        
        # Normalize column names
        if reader.fieldnames:
            normalized_fieldnames = []
            for col in reader.fieldnames:
                # Strip quotes, newlines, carriage returns, spaces
                clean_col = col.strip('"').strip().replace('\n', '').replace('\r', '').strip()
                normalized_fieldnames.append(clean_col)
            reader.fieldnames = normalized_fieldnames
        
        rows = list(reader)
        
        if len(rows) == 0:
            return False, "CSV file contains no data rows."
        
        # Clean all values in rows (remove quotes and extra whitespace)
        cleaned_rows = []
        for row in rows:
            cleaned_row = {}
            for key, value in row.items():
                if value is not None:
                    # Strip quotes and whitespace from values
                    cleaned_value = str(value).strip('"').strip()
                    cleaned_row[key] = cleaned_value
                else:
                    cleaned_row[key] = ''
            cleaned_rows.append(cleaned_row)
        
        return True, cleaned_rows
        
    except Exception as e:
        return False, f"Error parsing CSV: {str(e)}"


def validate_index_csv_columns(rows):
    """
    Validate that required columns exist in the Index CSV
    """
    if not rows:
        return False, "No data rows found."
    
    first_row = rows[0]
    
    # Get all column names for debugging
    available_columns = list(first_row.keys())
    
    # Check for required columns (with flexible matching)
    required_columns = ['INDEX', 'CURRENT', 'HIGH', 'LOW', 'PREV. CLOSE']
    
    found_columns = {}
    
    # Try to find each required column
    for req_col in required_columns:
        found = False
        for col_name in available_columns:
            # Normalize both for comparison
            col_normalized = col_name.upper().strip()
            req_normalized = req_col.upper().strip()
            
            # Check for exact match or if required column is contained in actual column
            if req_normalized == col_normalized or req_normalized in col_normalized:
                found_columns[req_col] = col_name
                found = True
                break
        
        if not found:
            return False, f"Missing required column: {req_col}. Available columns: {', '.join(available_columns)}"
    
    return True, found_columns


def filter_and_process_index_data(rows, nifty_symbols, column_map):
    """
    Filter rows by nifty symbols and process the index data
    Returns list of processed row dictionaries
    """
    processed_rows = []
    
    for row in rows:
        index_name = row.get(column_map['INDEX'], '').strip().strip('"').strip()
        
        # Filter by nifty symbols
        if index_name not in nifty_symbols:
            continue
        
        try:
            # Extract and convert numeric values (remove commas, quotes, and whitespace)
            def clean_numeric(value):
                """Clean and convert numeric string to float"""
                if not value:
                    return 0.0
                cleaned = str(value).replace(',', '').replace('"', '').strip()
                # Handle dash or empty values
                if cleaned == '-' or cleaned == '':
                    return 0.0
                return float(cleaned)
            
            close = clean_numeric(row.get(column_map['CURRENT'], '0'))
            high = clean_numeric(row.get(column_map['HIGH'], '0'))
            low = clean_numeric(row.get(column_map['LOW'], '0'))
            prv = clean_numeric(row.get(column_map['PREV. CLOSE'], '0'))
            
            # Calculate derived values
            # % = (CLOSE-PRV)*100/PRV
            percentage = ((close - prv) * 100 / prv) if prv != 0 else 0
            
            # PP = ABS((CLOSE+HIGH+LOW)/3)
            pp = abs((close + high + low) / 3)
            
            # R1 = ABS(PP*2-LOW)
            r1 = abs(pp * 2 - low)
            
            # S1 = ABS(PP*2-HIGH)
            s1 = abs(pp * 2 - high)
            
            # R2 = ABS(R1-S1+PP)
            r2 = abs(r1 - s1 + pp)
            
            # S2 = ABS(PP-(R1-S1))
            s2 = abs(pp - (r1 - s1))
            
            # R3 = ABS(HIGH+2*(PP-LOW))
            r3 = abs(high + 2 * (pp - low))
            
            # S3 = ABS(LOW-2*(HIGH-PP))
            s3 = abs(low - 2 * (high - pp))
            
            processed_rows.append({
                'INDEX': index_name,
                'CLOSE': close,
                'S1': s1,
                'S2': s2,
                'S3': s3,
                'PP': pp,
                'R1': r1,
                'R2': r2,
                'R3': r3,
                '%': percentage,
                'HIGH': high,
                'LOW': low,
                'PRV': prv
            })
            
        except (ValueError, TypeError) as e:
            # Skip rows with invalid numeric data
            continue
    
    return processed_rows


def create_index_excel_output(processed_rows, filename):
    """
    Create Excel file with processed index data
    Returns (success, filepath or error_message)
    """
    try:
        # Sort processed rows by % column in descending order
        processed_rows_sorted = sorted(processed_rows, key=lambda x: x.get('%', 0), reverse=True)
        
        # Create workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Index Chart"
        
        # Define column headers in the required order (no VOL, no SERIES)
        headers = ['INDEX', 'CLOSE', 'S1', 'S2', 'S3', 'PP', 
                  'R1', 'R2', 'R3', '%', 'HIGH', 'LOW', 'PRV']
        
        # Write headers with formatting
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows (sorted by % descending)
        for row_idx, data_row in enumerate(processed_rows_sorted, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = data_row.get(header, '')
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Format numeric cells to 2 decimal places (except INDEX)
                if header != 'INDEX' and isinstance(value, (int, float)):
                    cell.number_format = '0.00'
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            if header == 'INDEX':
                sheet.column_dimensions[chr(64 + col_idx)].width = 25
            else:
                sheet.column_dimensions[chr(64 + col_idx)].width = 12
        
        # Get Downloads folder path (cross-platform)
        downloads_folder = str(Path.home() / "Downloads")
        filepath = os.path.join(downloads_folder, filename)
        
        # Save workbook
        workbook.save(filepath)
        
        return True, filepath
        
    except Exception as e:
        return False, f"Error creating Excel file: {str(e)}"


def prepare_chart(chart_type, csv_content):
    """
    Main function to prepare chart from CSV data
    """
    try:
        # Handle Index Chart (market type)
        if chart_type == "market":
            # Step 1: Load nifty symbols
            success, result = load_nifty_symbols()
            if not success:
                return {
                    "success": False,
                    "message": result
                }
            nifty_symbols = result
            
            # Step 2: Parse Index CSV
            success, result = parse_index_csv(csv_content)
            if not success:
                return {
                    "success": False,
                    "message": result
                }
            csv_rows = result
            
            # Step 3: Validate required columns
            success, column_map = validate_index_csv_columns(csv_rows)
            if not success:
                return {
                    "success": False,
                    "message": column_map
                }
            
            # Step 4: Filter and process data
            processed_rows = filter_and_process_index_data(csv_rows, nifty_symbols, column_map)
            
            if not processed_rows:
                return {
                    "success": False,
                    "message": "No matching index symbols found in the CSV file."
                }
            
            # Step 5: Create Excel output
            date = datetime.now().strftime("%d-%m-%Y")
            filename = f"Index Chart {date}.xlsx"
            
            success, result = create_index_excel_output(processed_rows, filename)
            if not success:
                return {
                    "success": False,
                    "message": result
                }
            
            filepath = result
            
            return {
                "success": True,
                "message": f"Index Chart generated successfully with {len(processed_rows)} rows.",
                "filepath": filepath
            }
        
        # Handle Daily Chart
        if chart_type != "daily":
            return {
                "success": False,
                "message": "Unknown chart type."
            }
        
        # Step 1: Load stock symbols
        success, result = load_stock_symbols()
        if not success:
            return {
                "success": False,
                "message": result
            }
        stock_symbols = result
        
        # Step 2: Parse CSV with flexible column handling
        success, result = parse_csv_with_flexible_columns(csv_content)
        if not success:
            return {
                "success": False,
                "message": result
            }
        csv_rows = result
        
        # Step 3: Validate required columns
        success, message = validate_csv_columns(csv_rows)
        if not success:
            return {
                "success": False,
                "message": message
            }
        
        # Step 4: Filter and process data
        processed_rows = filter_and_process_data(csv_rows, stock_symbols)
        
        if not processed_rows:
            return {
                "success": False,
                "message": "No matching stock symbols found in the CSV file."
            }
        
        # Step 5: Create Excel output
        date = datetime.now().strftime("%d-%m-%Y")
        filename = f"Daily Chart {date}.xlsx"
        
        success, result = create_excel_output(processed_rows, filename)
        if not success:
            return {
                "success": False,
                "message": result
            }
        
        filepath = result
        
        return {
            "success": True,
            "message": f"Daily Chart generated successfully with {len(processed_rows)} rows.",
            "filepath": filepath
        }
        
    except Exception as ex:
        return {
            "success": False,
            "message": f"Unexpected error: {str(ex)}"
        }